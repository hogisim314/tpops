"""엑셀 export 관련 라우터"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

from database import get_db
from models import Server, Service, User, UserRole
from auth import get_current_active_user

router = APIRouter(prefix="/api/export", tags=["export"])


def map_db_to_display(db_info: str) -> str:
    """DB 정보를 표시용으로 변환"""
    if not db_info:
        return ""
    # DB:db_name 형식을 처리
    if db_info.startswith("DB:"):
        return db_info[3:]  # "DB:" 제거
    return db_info


@router.get("/servers")
async def export_servers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """서버 목록 엑셀 export (INFRASTRUCTURE 이상만)"""
    # 권한 체크
    if current_user.role not in [UserRole.INFRASTRUCTURE, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="권한이 없습니다. INFRASTRUCTURE 이상만 사용 가능합니다.")
    
    # 모든 서버 조회
    servers = db.query(Server).all()
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Servers"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="3a7bd5", end_color="3a7bd5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["서버명", "서버그룹", "노드", "MIN", "MAX", "재시작", "MAXQCOUNT", "ASQCOUNT", "DB연결"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_num, server in enumerate(servers, 2):
        ws.cell(row=row_num, column=1, value=server.name)
        ws.cell(row=row_num, column=2, value=server.svg_name)
        ws.cell(row=row_num, column=3, value=server.node_name)
        ws.cell(row=row_num, column=4, value=server.min_proc)
        ws.cell(row=row_num, column=5, value=server.max_proc)
        ws.cell(row=row_num, column=6, value=server.restart)
        ws.cell(row=row_num, column=7, value=server.maxqcount or "")
        ws.cell(row=row_num, column=8, value=server.asqcount or "")
        ws.cell(row=row_num, column=9, value=map_db_to_display(server.db_info) if server.db_info else "")
    
    # 컬럼 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 파일명 생성 (현재 날짜시간 포함)
    filename = f"servers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/services")
async def export_services(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """서비스 목록 엑셀 export (INFRASTRUCTURE 이상만)"""
    # 권한 체크
    if current_user.role not in [UserRole.INFRASTRUCTURE, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="권한이 없습니다. INFRASTRUCTURE 이상만 사용 가능합니다.")
    
    # 모든 서비스 조회
    services = db.query(Service).all()
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Services"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="3a7bd5", end_color="3a7bd5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["서비스명", "서버명", "타임아웃", "AutoTran", "Export"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_num, service in enumerate(services, 2):
        ws.cell(row=row_num, column=1, value=service.name)
        ws.cell(row=row_num, column=2, value=service.server_name)
        ws.cell(row=row_num, column=3, value=service.timeout)
        ws.cell(row=row_num, column=4, value=service.autotran)
        ws.cell(row=row_num, column=5, value=service.export)
    
    # 컬럼 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 파일명 생성 (현재 날짜시간 포함)
    filename = f"services_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetm l.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
